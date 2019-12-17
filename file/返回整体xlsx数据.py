#pip install openpyxl -->xlsx
'''
工作表是字典的键；
内容是字典的值；
值是一个列表，每一行又算一个列表，这些列表都是“值”这个大列表里的列表

'''
from openpyxl.reader.excel import load_workbook

def readExcel(path):
    dic = {}
    #打开文件
    file = load_workbook(filename=path)
    #所有工作表的名称
    print(file.get_sheet_names())
    sheets = file.get_sheet_names()

    for sheetName in sheets:
        sheet = file.get_sheet_by_name(sheetName)
        #一张表的所有数据
        sheetInfo = []
        for lineNum in range(1,sheet.max_row+1):
            #print(lineNum)
            lineList = []
            for columnNum in range(1,sheet.max_column+1):
                #拿数据
                value = sheet.cell(row=lineNum,column=columnNum).value
                #if value != None:
                lineList.append(value)

            sheetInfo.append(lineList)
        #将一张表的数据存到字典
        dic[sheetName] = sheetInfo
    return dic


#不能处理xls文件
path = ""
dic = readExcel(path)
print(len(dic))
print(dic['sheetname'])