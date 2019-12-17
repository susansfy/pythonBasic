#pip install openpyxl -->xlsx
'''
工作表是字典的键；
内容是字典的值；
值是一个列表，每一行又算一个列表，这些列表都是“值”这个大列表里的列表

'''
from openpyxl.reader.excel import load_workbook

def readExcel(path):
    #打开文件
    file = load_workbook(filename=path)
    #所有工作表的名称
    print(file.get_sheet_names())
    sheets = file.get_sheet_names()
    #拿出一个表格
    sheet = file.get_sheet_by_name(sheets[0])
    #最大行数
    #print(sheet.max_row)
    #最大列数
    #print(sheet.max_column)
    #表名
    #print(sheet.title)

    #表格的行数从1开始，
    for lineNum in range(1,sheet.max_row+1):
        #print(lineNum)
        lineList = []
        for columnNum in range(1,sheet.max_column+1):
            #拿数据
            value = sheet.cell(row=lineNum,column=columnNum).value
            #if value != None:
            lineList.append(value)
        #输出每行数据
        print(lineList)

#不能处理xls文件
path = ""
readExcel(path)